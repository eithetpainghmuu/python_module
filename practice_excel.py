from openpyxl import Workbook,load_workbook

#wb = Workbook()
#wb = load_workbook()

wb = load_workbook("python_excel.xlsx")
ws = wb["Sheet1"]

for row in ws.rows:
    print(row)
    for cell in row:
       ws['A1'] = 100



"""for row in ws.rows:
    #print(row)
    for cell in row:
        print(cell.value)

for i in range(4):
    print(i)
    for j in range(4):
     print(j)

for row in ws.iter_rows(min_row=1, min_col=1, max_row=4, max_col=2):
    for cell in row:
        print(cell.value)"""





#print(ws['A1'].value)
#print(ws['A2'].value)
#print(ws.cell(2,1).value)
#print(ws.cell(3,1).value)
#ws = wb.active
#ws = wb.create_sheet(title='newSheet')

#ws["A1"] = 1
#ws.cell(1,1,150)
#ws.cell(1,2,200)
#ws.cell(2,1,100)
#ws.cell(2,2,200)
#ws["B1"] = 2
#ws["A2"] = 3
#ws["B2"] = 4

#row_data = [1,2,3,4]
#ws.append(row_data)


#wb.save("test.xlsx")
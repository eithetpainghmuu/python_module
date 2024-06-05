import pyodbc
import openpyxl
# import sqlite3
# import csv
import pandas as pd
from openpyxl import Workbook
from dotenv import load_dotenv
import os

load_dotenv()

SERVER =os.getenv("SERVER")
DATABASE =os.getenv("DATABASE")
DB_USER =os.getenv("DB_USER")
DB_PASSWORD =os.getenv("DB_PASSWORD")

df = pd.read_csv("./mymoviedb.csv",lineterminator="\n")

conn = pyodbc.connect(f'DRIVER={{SQL Server}};SERVER={SERVER};DATABASE={DATABASE};UID={DB_USER};PWD={DB_PASSWORD}')

cursor = conn.cursor()


# for row in df.itertuples():
#     cursor.execute(
#         """
#         INSERT INTO movies (Release_Date, Title, Overview, Popularity, Vote_Count, Vote_Average, Original_Language, Genre, Poster_Url)
#         VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
#         """,
#         row.Release_Date,
#         row.Title,
#         row.Overview,
#         row.Popularity,
#         row.Vote_Count,
#         row.Vote_Average,
#         row.Original_Language,
#         row.Genre,
#         row.Poster_Url
#     )

# df = pd.read_sql("SELECT * FROM movies",conn)
#print(df)

gpLanguage = df.groupby("Original_Language")["Original_Language"].count()
# print(gpLanguage)

average = df.groupby("Genre")["Vote_Count"].mean()
#print(average)


# Create a new Workbook object
wb = Workbook()

ws = wb.active
ws.title = "Language Counts"
ws1 = wb.create_sheet(title = "Genre Ratings")

i = 1
for language,count in gpLanguage.items():
    ws.cell(i,1,language)
    ws.cell(i,2,count)
    i += 1

x = 1  
for Genre,Ratings in average.items():
    ws1.cell(x,1,Genre)
    ws1.cell(x,2,Ratings)
    x += 1

#Save the workbook to a file
wb.save('final_result.xlsx')
wb.close()

# Read data from Excel file
file_path = "final_result.xlsx"  # Update with your file path
df = pd.read_excel(file_path)

# Display the DataFrame
print(df)

# Commit the transaction
# conn.commit()

# Close the cursor and the database connection
cursor.close()
conn.close()